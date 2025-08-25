import os

import matplotlib as mpl
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import sympy as sp
from matplotlib import ticker
from scipy.interpolate import interp1d
from scipy.signal import savgol_filter
from sympy import Poly, Expr, fraction, expand, together, nsimplify
from sympy.polys.polyerrors import PolynomialError
# 配置字体支持
def setup_fonts():
    """Setup font support for English fonts only"""
    try:
        # Set global font configuration
        plt.rcParams['axes.unicode_minus'] = False  # Use ASCII minus instead of Unicode minus
        plt.rcParams['text.usetex'] = False
        plt.rcParams['mathtext.fontset'] = 'cm'
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        
        # English fonts only
        english_fonts = ['DejaVu Sans', 'Arial', 'Helvetica', 'Liberation Sans', 'FreeSans']
        # Check and select available fonts
        available_fonts = []
        for font_name in english_fonts:
            try:
                font_path = fm.findfont(fm.FontProperties(family=font_name))
                if font_path and os.path.exists(font_path):
                    available_fonts.append(font_name)
                    # Use DejaVu Sans as it handles minus signs well
                    if "DejaVu Sans" in font_path:
                        plt.rcParams['font.sans-serif'] = [font_name]
                        print(f"Using DejaVu Sans font for minus sign support")
                        return True
            except Exception as e:
                print(f"Error checking font {font_name}: {e}")
                continue

        # If available fonts found, use the first one
        if available_fonts:
            plt.rcParams['font.sans-serif'] = [available_fonts[0]]
            print(f"Using available font: {available_fonts[0]}")
            return True

        # Fallback: use DejaVu Sans if available
        try:
            dejavu_path = fm.findfont('DejaVu Sans')
            if dejavu_path:
                plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
                plt.rcParams['axes.unicode_minus'] = False
                print(f"Using DejaVu Sans font for minus sign support")
                return True
        except:
            pass

        # Final fallback: use system default font
        print("No suitable fonts found, using system default font")
        return False

    except Exception as e:
        print(f"Font setup error: {e}")
        return False


def safe_set_text(ax, xlabel=None, ylabel=None, title=None):
    """Safely set chart text, handle minus sign issues"""
    try:
        # Set font for all text properties
        for prop in [ax.xaxis.label, ax.yaxis.label, ax.title]:
            prop.set_family('DejaVu Sans')
            prop.set_size(10)  # Safe font size

        # Set labels normally
        if xlabel:
            ax.set_xlabel(xlabel)
        if ylabel:
            ax.set_ylabel(ylabel)
        if title:
            ax.set_title(title)

        # Force ASCII minus signs
        ax.xaxis.set_major_formatter(mpl.ticker.ScalarFormatter(useMathText=True))
        ax.yaxis.set_major_formatter(mpl.ticker.ScalarFormatter(useMathText=True))

        return True
    except:
        # If failed, return False
        return False


def setup_plot_formatting(ax):
    """Setup plot formatting, handle Unicode minus sign issues"""
    try:
        # Ensure ax is a valid Axes object
        if not hasattr(ax, 'xaxis'):
            return

        # Set number formatting to avoid Unicode minus sign issues
        import matplotlib.ticker as ticker

        # Safe method to handle scientific notation formatting
        def safe_sci_formatter(val, pos=None):
            """Safely format scientific notation numbers, handle minus sign issues"""
            s = f"{val:.1e}"
            s = s.replace('−', '-')  # Replace Unicode minus sign
            s = s.replace('e-', '×10⁻')  # Replace scientific notation
            s = s.replace('e+', '×10⁺')  # Replace scientific notation
            return s

        # Safe method to handle regular number formatting
        def safe_minus_formatter(val, pos=None):
            """Safely format regular numbers, handle minus sign issues"""
            s = f"{val:.2f}"
            return s.replace('−', '-')  # Replace Unicode minus sign

        # Create scientific notation and regular number formatters
        sci_formatter = ticker.FuncFormatter(lambda val, pos: safe_sci_formatter(val, pos))
        num_formatter = ticker.FuncFormatter(lambda val, pos: safe_minus_formatter(val, pos))

        # Set formatters for axes
        if hasattr(ax.xaxis, 'set_major_formatter'):
            # If value range is large, use scientific notation formatting
            x_range = np.ptp(ax.get_xlim())
            if x_range > 1e3:
                ax.xaxis.set_major_formatter(sci_formatter)
            else:
                ax.xaxis.set_major_formatter(num_formatter)

        if hasattr(ax.yaxis, 'set_major_formatter'):
            # If value range is large, use scientific notation formatting
            y_range = np.ptp(ax.get_ylim())
            if y_range > 1e3:
                ax.yaxis.set_major_formatter(sci_formatter)
            else:
                ax.yaxis.set_major_formatter(num_formatter)

        # Set tick label format
        if hasattr(ax, 'tick_params'):
            ax.tick_params(axis='both', which='major', labelsize=10)
            ax.tick_params(axis='both', which='minor', labelsize=8)

    except Exception as e:
        print(f"Plot formatting setup error: {e}")
        # If failed, fallback to global settings
        plt.rcParams['axes.unicode_minus'] = False


class ThermalAnalysisProcessor:
    def __init__(self, precision='float64'):
        """
        初始化热分析处理器
        
        Args:
            precision (str): 计算精度，可选 'float32' 或 'float64'，默认为 'float64'
        """
        self.t0 = None
        self.Tj = None
        self.Zth_transient = None
        self.ploss = 1.0
        self.delta_z = 0.05
        self.num_iterations = 500
        self.discrete_order = 35
        self.results = {}

        # 设置计算精度
        self.precision = precision
        if precision == 'float32':
            self.dtype = np.float32
        else:
            self.dtype = np.float64

        print(f"热分析处理器初始化完成，计算精度: {precision}")

    def load_data(self, file_path):
        """从Excel文件加载数据 - 支持多种格式"""
        try:
            # 方法1: 使用openpyxl引擎读取Excel
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
            except:
                # 方法2: 使用xlrd引擎读取旧版Excel
                try:
                    df = pd.read_excel(file_path, engine='xlrd')
                except:
                    # 方法3: 使用csv格式读取
                    if file_path.endswith('.csv'):
                        df = pd.read_csv(file_path)
                    else:
                        # 方法4: 使用numpy直接读取文本文件
                        data = np.loadtxt(file_path, delimiter=',', skiprows=1)
                        df = pd.DataFrame(data, columns=['time', 'temperature'])

            # 提取数据
            if len(df.columns) >= 2:
                self.t0 = df.iloc[:, 0].values.astype(self.dtype)
                self.Tj = df.iloc[:, 1].values.astype(self.dtype)
            else:
                raise ValueError("数据文件至少需要两列：时间和温度")

            # 数据验证
            if len(self.t0) == 0 or len(self.Tj) == 0:
                raise ValueError("数据为空")

            if len(self.t0) != len(self.Tj):
                raise ValueError("时间和温度数据长度不匹配")

            # 数据清理：移除无效值
            valid_mask = (self.t0 > 0) & (np.isfinite(self.t0)) & (np.isfinite(self.Tj))
            if not np.all(valid_mask):
                print(f"警告: 发现 {np.sum(~valid_mask)} 个无效数据点，已自动移除")
                self.t0 = self.t0[valid_mask]
                self.Tj = self.Tj[valid_mask]

            # 检查重复时间值
            unique_times, unique_indices = np.unique(self.t0, return_index=True)
            if len(unique_times) != len(self.t0):
                print(f"警告: 发现 {len(self.t0) - len(unique_times)} 个重复时间值，已自动处理")
                # 保留第一个出现的值
                unique_indices = np.sort(unique_indices)
                self.t0 = self.t0[unique_indices]
                self.Tj = self.Tj[unique_indices]

            print(f"成功加载数据: {len(self.t0)} 个有效数据点")
            return True

        except Exception as e:
            print(f"Error loading data: {e}")
            return False

    def load_time_constant_spectrum_data(self, file_path):
        """从Excel文件直接加载时间常数谱数据 - 第一列为时间，第二列为R(z)"""
        try:
            # 方法1: 使用openpyxl引擎读取Excel
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
            except:
                # 方法2: 使用xlrd引擎读取旧版Excel
                try:
                    df = pd.read_excel(file_path, engine='xlrd')
                except:
                    # 方法3: 使用csv格式读取
                    if file_path.endswith('.csv'):
                        df = pd.read_csv(file_path)
                    else:
                        # 方法4: 使用numpy直接读取文本文件
                        data = np.loadtxt(file_path, delimiter=',', skiprows=1)
                        df = pd.DataFrame(data, columns=['time', 'R(z)'])

            # 提取数据
            if len(df.columns) >= 2:
                t_bayesian = df.iloc[:, 0].values.astype(self.dtype)
                R = df.iloc[:, 1].values.astype(self.dtype)
            else:
                raise ValueError("数据文件至少需要两列：时间和时间常数谱R(z)")

            # 数据验证
            if len(t_bayesian) == 0 or len(R) == 0:
                raise ValueError("数据为空")

            if len(t_bayesian) != len(R):
                raise ValueError("时间和时间常数谱数据长度不匹配")

            # 数据清理：移除无效值
            valid_mask = (t_bayesian > 0) & (np.isfinite(t_bayesian)) & (np.isfinite(R))
            if not np.all(valid_mask):
                print(f"警告: 发现 {np.sum(~valid_mask)} 个无效数据点，已自动移除")
                t_bayesian = t_bayesian[valid_mask]
                R = R[valid_mask]

            # 检查重复时间值
            unique_times, unique_indices = np.unique(t_bayesian, return_index=True)
            if len(unique_times) != len(t_bayesian):
                print(f"警告: 发现 {len(t_bayesian) - len(unique_times)} 个重复时间值，已自动处理")
                # 保留第一个出现的值
                unique_indices = np.sort(unique_indices)
                t_bayesian = t_bayesian[unique_indices]
                R = R[unique_indices]

            # 计算对应的z_bayesian
            z_bayesian = np.log(t_bayesian).astype(self.dtype)

            # 保存到results中，模拟bayesian_deconvolution的结果
            self.results['t_bayesian'] = t_bayesian
            self.results['z_bayesian'] = z_bayesian
            self.results['R'] = R
            self.results['z_bayesian_for_R'] = z_bayesian

            print(f"成功加载时间常数谱数据: {len(t_bayesian)} 个有效数据点")
            print(f"时间范围: {t_bayesian.min():.2e} - {t_bayesian.max():.2e} s")
            print(f"R(z)范围: {R.min():.2e} - {R.max():.2e}")
            return True

        except Exception as e:
            print(f"Error loading time constant spectrum data: {e}")
            return False

    def load_data_alternative(self, file_path):
        """从文件加载数据的替代方法 - 使用不同的库"""
        try:
            import csv
            import io

            # 方法1: 使用csv模块读取
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as file:
                    csv_reader = csv.reader(file)
                    next(csv_reader)  # 跳过标题行
                    data = list(csv_reader)

                if len(data) > 0 and len(data[0]) >= 2:
                    self.t0 = np.array([float(row[0]) for row in data], dtype=self.dtype)
                    self.Tj = np.array([float(row[1]) for row in data], dtype=self.dtype)
                    print(f"使用csv模块加载数据: {len(self.t0)} 个数据点")
                    return True

            # 方法2: 使用openpyxl直接读取Excel
            elif file_path.endswith(('.xlsx', '.xls')):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active

                    data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
                        if row[0] is not None and row[1] is not None:
                            data.append([float(row[0]), float(row[1])])

                    if len(data) > 0:
                        self.t0 = np.array([row[0] for row in data], dtype=self.dtype)
                        self.Tj = np.array([row[1] for row in data], dtype=self.dtype)
                        print(f"使用openpyxl加载数据: {len(self.t0)} 个数据点")
                        return True

                except ImportError:
                    print("openpyxl未安装，尝试其他方法")

            # 方法3: 使用xlrd直接读取旧版Excel
            elif file_path.endswith('.xls'):
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(file_path)
                    sheet = workbook.sheet_by_index(0)

                    data = []
                    for row_idx in range(1, sheet.nrows):  # 跳过标题行
                        row = sheet.row_values(row_idx)
                        if row[0] and row[1]:
                            data.append([float(row[0]), float(row[1])])

                    if len(data) > 0:
                        self.t0 = np.array([row[0] for row in data], dtype=self.dtype)
                        self.Tj = np.array([row[1] for row in data], dtype=self.dtype)
                        print(f"使用xlrd加载数据: {len(self.t0)} 个数据点")
                        return True

                except ImportError:
                    print("xlrd未安装，尝试其他方法")

            # 方法4: 使用numpy读取文本文件
            else:
                try:
                    data = np.genfromtxt(file_path, delimiter=',', skip_header=1)
                    if data.shape[1] >= 2:
                        self.t0 = data[:, 0].astype(self.dtype)
                        self.Tj = data[:, 1].astype(self.dtype)
                        print(f"使用numpy加载数据: {len(self.t0)} 个数据点")
                        return True
                except:
                    pass

            raise ValueError("无法读取文件格式")

        except Exception as e:
            print(f"Alternative loading error: {e}")
            return False

    def calculate_zth(self, ambient_temp=25.0):
        """计算瞬态热阻抗"""
        if self.t0 is None or self.Tj is None:
            return False

        self.Zth_transient = ((self.Tj - ambient_temp) / self.ploss).astype(self.dtype)
        return True

    def logarithmic_interpolation(self):
        """对时间轴取对数并进行插值"""
        if self.t0 is None or self.Zth_transient is None:
            return False

        # 数据预处理：移除零值和负值，处理重复值
        valid_mask = (self.t0 > 0) & (self.Zth_transient > -np.inf) & (self.Zth_transient < np.inf)
        t0_clean = self.t0[valid_mask]
        Zth_clean = self.Zth_transient[valid_mask]

        if len(t0_clean) < 2:
            print("Error: 没有有效的时间数据点")
            return False

        # 移除重复的时间值（保留第一个出现的值）
        unique_indices = np.unique(t0_clean, return_index=True)[1]
        unique_indices = np.sort(unique_indices)
        t0_unique = t0_clean[unique_indices]
        Zth_unique = Zth_clean[unique_indices]

        if len(t0_unique) < 2:
            print("Error: 去除重复值后数据点不足")
            return False

        # 步骤1: 时间轴取对数（现在安全了）
        z_fft = np.log(t0_unique)

        # 步骤2: 创建插值函数（使用线性插值避免重复值问题）
        try:
            interp_func = interp1d(t0_unique, Zth_unique, kind='linear', fill_value='extrapolate')
        except:
            # 如果还有问题，使用最简单的线性插值
            interp_func = interp1d(t0_unique, Zth_unique, kind='linear',
                                   fill_value=(Zth_unique[0], Zth_unique[-1]),
                                   bounds_error=False)

        # 步骤3: 计算 a(z) = Zth(t=exp(z))
        az_fft = interp_func(np.exp(z_fft))

        # 步骤4: 均匀插值
        z_end = np.log(t0_unique[-1])
        z_bayesian = np.arange(-9.2, z_end + self.delta_z, self.delta_z, dtype=self.dtype)
        t_bayesian = np.exp(z_bayesian).astype(self.dtype)

        # 确保插值范围在有效数据范围内
        t_bayesian = t_bayesian[(t_bayesian >= t0_unique[0]) & (t_bayesian <= t0_unique[-1])]
        az_bayesian = interp_func(t_bayesian).astype(self.dtype)
        z_bayesian = np.log(t_bayesian).astype(self.dtype)

        self.results['z_fft'] = z_fft
        self.results['az_fft'] = az_fft
        self.results['z_bayesian'] = z_bayesian
        self.results['t_bayesian'] = t_bayesian
        self.results['az_bayesian'] = az_bayesian

        print(f"对数插值完成: 原始数据点 {len(self.t0)}, 有效数据点 {len(t0_unique)}, 插值数据点 {len(t_bayesian)}")
        return True

    def calculate_derivative(self):
        """计算导数 da(z)/dz"""
        if 'z_bayesian' not in self.results or 'az_bayesian' not in self.results:
            return False

        z_bayesian = self.results['z_bayesian']
        az_bayesian = self.results['az_bayesian']

        # 计算导数
        dz_bayesian = np.diff(z_bayesian).astype(self.dtype)
        da_bayesian = np.diff(az_bayesian).astype(self.dtype)
        da_dz_bayesian = (da_bayesian / dz_bayesian).astype(self.dtype)

        # 平滑导数
        window_size = min(51, len(da_dz_bayesian))
        if window_size % 2 == 0:
            window_size -= 1  # 确保窗口大小为奇数
        da_dz_bayesian_smoothed = savgol_filter(da_dz_bayesian, window_size, 3).astype(self.dtype)

        self.results['dz_bayesian'] = dz_bayesian
        self.results['da_dz_bayesian'] = da_dz_bayesian
        self.results['da_dz_bayesian_smoothed'] = da_dz_bayesian_smoothed

        return True

    def calculate_weight_function(self):
        """计算权重函数 w(z) = exp(z - exp(z))"""
        if 'z_bayesian' not in self.results:
            return False

        z_bayesian = self.results['z_bayesian']
        wz_bayesian = np.exp(z_bayesian - np.exp(z_bayesian)).astype(self.dtype)

        self.results['wz_bayesian'] = wz_bayesian
        return True

    def bayesian_deconvolution(self):
        """贝叶斯反卷积计算时间常数谱"""
        if ('z_bayesian' not in self.results or
                'da_dz_bayesian_smoothed' not in self.results or
                'wz_bayesian' not in self.results):
            return False

        z_bayesian = self.results['z_bayesian']
        da_dz_bayesian = self.results['da_dz_bayesian_smoothed']
        wz_bayesian = self.results['wz_bayesian']

        # 确保数组长度匹配
        if len(da_dz_bayesian) != len(z_bayesian) - 1:
            print(f"警告: da_dz_bayesian长度({len(da_dz_bayesian)})与预期长度({len(z_bayesian) - 1})不匹配")
            # 调整数组长度
            min_len = min(len(da_dz_bayesian), len(z_bayesian) - 1)
            da_dz_bayesian = da_dz_bayesian[:min_len]
            z_bayesian = z_bayesian[:min_len + 1]
            wz_bayesian = wz_bayesian[:min_len + 1]
            print(f"调整数组长度到: da_dz_bayesian={len(da_dz_bayesian)}, z_bayesian={len(z_bayesian)}")

        # 生成权重函数矩阵
        n = len(da_dz_bayesian)
        W = np.zeros((n, n), dtype=self.dtype)

        for i in range(n):
            for j in range(n):
                W[i, j] = np.exp(z_bayesian[i] - z_bayesian[j] - np.exp(z_bayesian[i] - z_bayesian[j]))

        # 初始化R
        R = (da_dz_bayesian / np.sum(da_dz_bayesian)).astype(self.dtype)

        # 迭代计算
        for iter in range(self.num_iterations):
            conv_w_R = W @ R
            xcorr_w_right = W.T @ (da_dz_bayesian / conv_w_R)
            R = R * xcorr_w_right

        self.results['R'] = R
        # 同时保存对应的z_bayesian（用于后续计算）
        self.results['z_bayesian_for_R'] = z_bayesian[:n]

        return True

    def discrete_time_constant_spectrum(self):
        """时间常数谱离散化 - 改进版本"""
        if ('z_bayesian' not in self.results or
                'R' not in self.results):
            return False

        # 优先使用与R对应的z_bayesian数组
        if 'z_bayesian_for_R' in self.results:
            z_bayesian = self.results['z_bayesian_for_R']
        else:
            z_bayesian = self.results['z_bayesian']

        R = self.results['R']

        # 检查数组长度匹配
        if len(R) != len(z_bayesian):
            print(f"警告: R数组长度({len(R)})与z_bayesian数组长度({len(z_bayesian)})不匹配")
            # 如果R数组比z_bayesian短1，这是正常的（因为R基于导数计算）
            if len(R) == len(z_bayesian) - 1:
                # 使用z_bayesian的前len(R)个元素
                z_bayesian = z_bayesian[:len(R)]
                print(f"调整z_bayesian数组长度以匹配R数组")
            else:
                print("错误: 无法处理数组长度不匹配")
                return False

        # 过滤掉无效的R值
        valid_mask = (R > 0) & np.isfinite(R)
        if not np.any(valid_mask):
            print("警告: 没有有效的时间常数谱数据")
            return False

        z_valid = z_bayesian[valid_mask]
        R_valid = R[valid_mask]

        # 改进：使用更智能的区间划分方法
        # 基于R值的分布进行自适应区间划分
        R_cumsum = np.cumsum(R_valid)
        R_total = R_cumsum[-1]

        # 确保每个区间包含足够的R值
        min_R_per_interval = R_total / (self.discrete_order * 2)  # 至少包含总R值的1/(2*order)

        # 创建区间边界
        interval_boundaries = []
        current_sum = 0
        target_sum = R_total / self.discrete_order

        for i, r_val in enumerate(R_valid):
            current_sum += r_val
            if current_sum >= target_sum and len(interval_boundaries) < self.discrete_order - 1:
                interval_boundaries.append(i)
                current_sum = 0

        # 确保有足够的区间
        if len(interval_boundaries) < self.discrete_order - 1:
            # 如果区间不够，使用均匀划分
            step = len(R_valid) // self.discrete_order
            interval_boundaries = [i * step for i in range(1, self.discrete_order)]

        # 添加起始和结束边界
        interval_boundaries = [0] + interval_boundaries + [len(R_valid)]

        # Foster网络参数计算
        fosterRth = []
        fosterCth = []
        tau_Foster = []

        for i in range(len(interval_boundaries) - 1):
            start_idx = interval_boundaries[i]
            end_idx = interval_boundaries[i + 1]

            if end_idx > start_idx:
                # 计算该区间内的总热阻
                interval_R = np.sum(R_valid[start_idx:end_idx]) * self.delta_z

                if interval_R > 0:
                    # 计算该区间的平均时间常数
                    z_interval = z_valid[start_idx:end_idx]
                    tau_interval = np.exp(z_interval)

                    # 使用加权平均计算时间常数
                    weights = R_valid[start_idx:end_idx]
                    tau_weighted = np.average(tau_interval, weights=weights)

                    # 计算对应的热容
                    interval_C = tau_weighted / interval_R

                    # 验证热容的合理性
                    if interval_C > 0 and np.isfinite(interval_C):
                        fosterRth.append(interval_R)
                        fosterCth.append(interval_C)
                        tau_Foster.append(interval_R * interval_C)

        # 转换为numpy数组
        fosterRth = np.array(fosterRth, dtype=self.dtype)
        fosterCth = np.array(fosterCth, dtype=self.dtype)
        tau_Foster = np.array(tau_Foster, dtype=self.dtype)

        # 过滤掉零值和负值
        valid_foster_mask = (fosterRth > 0) & (fosterCth > 0) & np.isfinite(fosterRth) & np.isfinite(fosterCth)
        if not np.any(valid_foster_mask):
            print("警告: 没有生成有效的Foster网络参数")
            return False

        self.results['fosterRth'] = fosterRth[valid_foster_mask]
        self.results['fosterCth'] = fosterCth[valid_foster_mask]
        self.results['tau_Foster'] = tau_Foster[valid_foster_mask]

        print(f"改进的Foster网络参数计算完成: {np.sum(valid_foster_mask)} 个有效参数")
        print(f"热阻范围: {fosterRth[valid_foster_mask].min():.2e} - {fosterRth[valid_foster_mask].max():.2e} K/W")
        print(f"热容范围: {fosterCth[valid_foster_mask].min():.2e} - {fosterCth[valid_foster_mask].max():.2e} Ws/K")
        return True

    def calc_parallel_c(self, s, numy, deny):
        """
        计算并联热容参数（等效于Matlab的calcParallelC函数）
        参数：
        s: 符号变量
        numY: 分子多项式表达式
        denY: 分母多项式表达式
        返回：
        c: 首项系数比值
        """
        try:
            # 将表达式转换为多项式并提取首项系数
            num_poly = Poly(numy, s)
            den_poly = Poly(deny, s)

            # 检查多项式是否为空或系数是否为空
            num_coeffs = num_poly.coeffs()
            den_coeffs = den_poly.coeffs()

            if len(num_coeffs) == 0 or len(den_coeffs) == 0:
                print("警告: 多项式系数为空")
                return 0
            # 获取最高次项的系数
            leading_coeff_num = num_coeffs[0]
            leading_coeff_den = den_coeffs[0]
            # 检查分母系数是否为零
            if leading_coeff_den == 0:
                print("警告: 分母系数为零")
                return 0
            return leading_coeff_num / leading_coeff_den
        except Exception as e:
            print(f"calc_parallel_c 计算错误: {e}")
            print(f"numy: {numy}")
            print(f"deny: {deny}")
            return 0

    def calc_series_r(self, s, numz, denz):
        """
        计算串联热阻参数（等效于Matlab的calcSeriesR函数）

        参数：
        s: 符号变量
        numZ: 分子多项式表达式
        denZ: 分母多项式表达式

        返回：
        r: 首项系数比值
        """
        try:
            # 将表达式转换为多项式并提取首项系数
            num_poly = Poly(numz, s)
            den_poly = Poly(denz, s)

            # 检查多项式是否为空或系数是否为空
            num_coeffs = num_poly.coeffs()
            den_coeffs = den_poly.coeffs()

            if len(num_coeffs) == 0 or len(den_coeffs) == 0:
                print("警告: 多项式系数为空")
                return 0

            # 获取最高次项的系数
            leading_coeff_num = num_coeffs[0]
            leading_coeff_den = den_coeffs[0]

            # 检查分母系数是否为零
            if leading_coeff_den == 0:
                print("警告: 分母系数为零")
                return 0

            return leading_coeff_num / leading_coeff_den

        except Exception as e:
            print(f"calc_series_r 计算错误: {e}")
            print(f"numz: {numz}")
            print(f"denz: {denz}")
            return 0

    def manual_handler(self):
        """手动调整Foster网络参数，确保参数合理且不会导致计算问题"""
        foster_rth = self.results['fosterRth'].copy()  # 创建副本避免修改原始数据
        foster_cth = self.results['fosterCth'].copy()

        # 确保参数不为零或过小，避免数值计算问题
        min_rth = 1e-6  # 最小热阻值
        min_cth = 1e-6  # 最小热容值

        # 调整前几个参数，但保持合理的比例关系
        if len(foster_rth) >= 3:
            foster_rth[0] = max(foster_rth[0], min_rth)
            foster_rth[1] = max(foster_rth[1], min_rth)
            foster_rth[2] = max(foster_rth[2], min_rth)

        if len(foster_cth) >= 6:
            foster_cth[0] = max(foster_cth[0], min_cth)
            foster_cth[1] = max(foster_cth[1], min_cth)
            foster_cth[2] = max(foster_cth[2], min_cth)
            foster_cth[3] = max(foster_cth[3], min_cth)
            foster_cth[4] = max(foster_cth[4], min_cth)
            foster_cth[5] = max(foster_cth[5], min_cth)

        # 确保所有参数都是有限的正数
        foster_rth = np.where((foster_rth > 0) & np.isfinite(foster_rth), foster_rth, min_rth)
        foster_cth = np.where((foster_cth > 0) & np.isfinite(foster_cth), foster_cth, min_cth)

        self.results['fosterRth'] = foster_rth
        self.results['fosterCth'] = foster_cth

        print(f"手动调整完成: 热阻范围 {foster_rth.min():.2e} - {foster_rth.max():.2e} K/W")
        print(f"热容范围 {foster_cth.min():.2e} - {foster_cth.max():.2e} Ws/K")

    def foster_to_cauer(self):
        """根据Matlab代码逻辑重写的Foster到Cauer网络转换方法 - 使用最大公因子辗转相除法"""
        foster_rth = self.results['fosterRth']
        foster_cth = self.results['fosterCth']
        s = sp.symbols('s')

        # 计算复阻抗 Z(s)
        Zs = 0
        for i in range(len(foster_rth)):
            Zs += foster_rth[i] / (1 + foster_rth[i] * foster_cth[i] * s)

        # 计算复导纳 Y(s)
        Ys = 1 / Zs

        # 提取 Y(s) 的分子和分母多项式
        try:
            numY, denY = self.normalized_numden(Ys, s, tol=1e-50)
            numY = sp.collect(numY, s)  # 简化分子
            denY = sp.collect(denY, s)  # 简化分母
        except Exception as e:
            print(f"提取分子分母多项式时出错: {e}")
            return False

        # 初始化参数
        cauerRth = []
        cauerCth = []

        j = 0
        max_iterations = len(foster_rth)
        
        for i in range(max_iterations):
            try:
                # 计算并联热容参数
                c = self.calc_parallel_c(s, numY, denY)
                if c == 0 or not sp.simplify(c).is_finite:
                    print(f"迭代 {i+1}: 并联热容参数无效，停止迭代")
                    break
                cauerCth.append(c)
                print(f"迭代 {i+1}: 并联热容 = {c}")
                
                # 更新复导纳的分子
                numY = sp.expand(numY - c * s * denY)
                if sp.expand(numY) == 0:
                    print(f"迭代 {i+1}: 更新后的分子为零，停止迭代")
                    break
                # 计算更新后的复阻抗 Z(s)
                try:
                    Zs = 1 / (numY / denY)
                except Exception as e:
                    print(f"迭代 {i + 1}: 计算Z(s)时出错: {e}")
                    break

                # 提取Z(s)的分子和分母多项式
                try:
                    numZ, denZ = self.normalized_numden(Zs, s, tol=1e-50)
                    numZ = sp.collect(numZ, s)  # 简化分子
                    denZ = sp.collect(denZ, s)  # 简化分母
                except Exception as e:
                    print(f"迭代 {i + 1}: 提取Z(s)分子分母时出错: {e}")
                    break

                # 计算串联热阻参数
                r = self.calc_series_r(s, numZ, denZ)
                if r == 0 or not sp.simplify(r).is_finite:
                    print(f"迭代 {i+1}: 串联热阻参数无效，停止迭代")
                    break
                cauerRth.append(r)
                print(f"迭代 {i+1}: 串联热阻 = {r}")
                
                # 更新复阻抗的分子
                numZ = sp.expand(numZ - r * denZ)

                j = j + 1
                numY = denZ
                denY = numZ

                # 判断 Z(s) 是否等于 0
                if sp.expand(numZ / denZ) == 0:
                    print(f"迭代 {i+1}: Z(s) = 0，停止迭代")
                    break
                    
            except Exception as e:
                print(f"迭代 {i+1} 时出错: {e}")
                break

        # 将符号表达式转换为数值
        try:
            cauerRth_numeric = []
            cauerCth_numeric = []

            for r_val in cauerRth:
                try:
                    # 尝试将符号表达式转换为浮点数
                    r_numeric = float(r_val)
                    cauerRth_numeric.append(r_numeric)
                except (TypeError, ValueError):
                    # 如果转换失败，跳过这个值
                    print(f"警告: 无法转换热阻值 {r_val} 为数值")
                    continue

            for c_val in cauerCth:
                try:
                    # 尝试将符号表达式转换为浮点数
                    c_numeric = float(c_val)
                    cauerCth_numeric.append(c_numeric)
                except (TypeError, ValueError):
                    # 如果转换失败，跳过这个值
                    print(f"警告: 无法转换热容值 {c_val} 为数值")
                    continue

            # 转换为numpy数组
            cauerRth_array = np.array(cauerRth_numeric, dtype=self.dtype)
            cauerCth_array = np.array(cauerCth_numeric, dtype=self.dtype)

            # 保存数值数组
            self.results['cauerRth'] = cauerRth_array
            self.results['cauerCth'] = cauerCth_array

            print(f"Foster到Cauer转换完成: {len(cauerRth_array)} 个有效参数")
            return True

        except Exception as e:
            print(f"Foster到Cauer转换失败: {e}")
            # 如果转换失败，保存原始符号表达式（用于调试）
            self.results['cauerRth'] = cauerRth
            self.results['cauerCth'] = cauerCth
            return False

    @staticmethod
    def normalized_numden(expr: Expr, s, tol=1e-50) -> (Expr, Expr):
        """
        规范化提取分子分母，特别处理高次小系数多项式

        参数：
        expr: SymPy表达式
        s: 符号变量
        tol: 系数截断阈值

        返回：
        (num, den): 分子和分母表达式
        """
        try:
            # 1. 数值稳定化处理
            expr = nsimplify(expr, rational=True)  # 尝试转换为有理数
            expr = expr.evalf()  # 转换为浮点数表达式

            # 2. 展开并规范化表达式
            expr = expand(expr)
            expr = together(expr)

            # 3. 安全提取分子分母
            if expr.is_rational_function():
                num, den = fraction(expr)
            else:
                num = expr
                den = sp.Integer(1)

            # 4. 系数截断处理
            def truncate_coeffs(poly_expr, var):
                """截断极小系数项"""
                if not poly_expr.is_polynomial(var):
                    return poly_expr

                # 转换为多项式并处理系数
                poly = Poly(poly_expr, var)
                coeffs = poly.all_coeffs()

                # 截断极小系数
                truncated_coeffs = [
                    coef if abs(coef) > tol else 0
                    for coef in coeffs
                ]

                # 重建多项式
                return sum(
                    c * var ** i
                    for i, c in enumerate(reversed(truncated_coeffs))
                    if c != 0
                )

            # 应用截断
            num = truncate_coeffs(num, s)
            den = truncate_coeffs(den, s)

            # 5. 对分母进行首项系数规范化
            try:
                den_poly = Poly(den, s)
                if den_poly.is_zero:
                    raise ValueError("分母不能为零")

                lc = den_poly.LC()
                if lc != 1 and lc != 0:
                    num = num / lc
                    den = den / lc
            except PolynomialError:
                # 非多项式分母，跳过规范化
                pass

            return num, den

        except Exception as e:
            raise RuntimeError(f"提取分子分母时出错: {str(e)}\n表达式: {expr}") from e

    def calculate_structure_functions(self):
        """计算结构函数 - 基于Cauer网络参数
        
        注意：结构函数应该基于Cauer网络参数而不是Foster网络参数，因为：
        1. Cauer网络是梯形网络，直接对应物理结构
        2. Cauer网络的热阻和热容按物理层次排列，适合计算结构函数
        3. 积分结构函数表示从热源到环境的总热阻和热容
        4. 微分结构函数表示每个物理层的热阻和热容分布
        
        计算逻辑基于Matlab代码：
        - 积分结构函数：cumsum(cauerRth) 和 cumsum(cauerCth)
        - 微分结构函数：diff(cumulative_Cth) / diff(cumulative_Rth)
        - 自动去除热容过大的异常层
        """
        if ('cauerRth' not in self.results or
                'cauerCth' not in self.results):
            print("警告: 没有Cauer网络参数，请先执行Foster到Cauer转换")
            return False

        cauerRth = self.results['cauerRth']
        cauerCth = self.results['cauerCth']

        # 确保转换为numpy数组
        try:
            if isinstance(cauerRth, list):
                # 如果是列表，尝试转换为数值数组
                cauerRth_numeric = []
                for r_val in cauerRth:
                    try:
                        r_numeric = float(r_val)
                        cauerRth_numeric.append(r_numeric)
                    except (TypeError, ValueError):
                        print(f"警告: 无法转换热阻值 {r_val} 为数值")
                        continue
                cauerRth = np.array(cauerRth_numeric, dtype=self.dtype)
            else:
                cauerRth = np.array(cauerRth, dtype=self.dtype)

            if isinstance(cauerCth, list):
                # 如果是列表，尝试转换为数值数组
                cauerCth_numeric = []
                for c_val in cauerCth:
                    try:
                        c_numeric = float(c_val)
                        cauerCth_numeric.append(c_numeric)
                    except (TypeError, ValueError):
                        print(f"警告: 无法转换热容值 {c_val} 为数值")
                        continue
                cauerCth = np.array(cauerCth_numeric, dtype=self.dtype)
            else:
                cauerCth = np.array(cauerCth, dtype=self.dtype)

        except Exception as e:
            print(f"警告: Cauer参数转换为数值失败: {e}")
            return False

        # 检查数组长度是否匹配
        if len(cauerRth) != len(cauerCth):
            print(f"警告: Cauer热阻数组长度({len(cauerRth)})与热容数组长度({len(cauerCth)})不匹配")
            # 使用较短的长度
            min_len = min(len(cauerRth), len(cauerCth))
            cauerRth = cauerRth[:min_len]
            cauerCth = cauerCth[:min_len]

        # 过滤掉零值、负值和无效值
        valid_mask = (cauerRth > 0) & (cauerCth > 0) & np.isfinite(cauerRth) & np.isfinite(cauerCth)
        if not np.any(valid_mask):
            print("警告: 没有有效的Cauer网络参数用于结构函数计算")
            return False

        cauerRth_valid = cauerRth[valid_mask]
        cauerCth_valid = cauerCth[valid_mask]

        # Cauer网络的结构函数计算 - 基于Matlab代码逻辑
        # 积分结构函数计算
        cumulative_Rth = np.cumsum(cauerRth_valid).astype(self.dtype)
        cumulative_Cth = np.cumsum(cauerCth_valid).astype(self.dtype)

        # 计算微分结构函数 - 修复数组长度问题
        if len(cumulative_Rth) > 1:
            # 计算相邻积分值的差值
            diff_cumulative_Rth = np.diff(cumulative_Rth)
            diff_cumulative_Cth = np.diff(cumulative_Cth)

            # 计算微分结构函数
            # 注意：避免除零错误
            valid_diff_mask = diff_cumulative_Rth > 0
            if np.any(valid_diff_mask):
                # 使用正确的数组长度
                differential_Cth = np.zeros(len(diff_cumulative_Rth))
                differential_Rth = np.zeros(len(diff_cumulative_Rth))

                # 对于有效差值，计算微分热容
                differential_Cth[valid_diff_mask] = diff_cumulative_Cth[valid_diff_mask] / diff_cumulative_Rth[
                    valid_diff_mask]
                differential_Rth[valid_diff_mask] = cumulative_Rth[:-1][valid_diff_mask]  # 使用前n-1个积分热阻值

                # 过滤掉无效值
                valid_differential_mask = (differential_Rth > 0) & (differential_Cth > 0) & np.isfinite(
                    differential_Rth) & np.isfinite(differential_Cth)
                if np.any(valid_differential_mask):
                    differential_Rth = differential_Rth[valid_differential_mask]
                    differential_Cth = differential_Cth[valid_differential_mask]
                else:
                    # 如果没有有效微分值，使用原始Cauer参数作为备选
                    differential_Rth = cauerRth_valid
                    differential_Cth = cauerCth_valid
            else:
                # 如果没有有效差值，使用原始Cauer参数
                differential_Rth = cauerRth_valid
                differential_Cth = cauerCth_valid
        else:
            # 如果只有一个数据点，则没有微分值
            differential_Rth = np.array([])
            differential_Cth = np.array([])

        # 去掉最后几个热容特别大的层（如果存在）
        if len(cumulative_Cth) > 4:
            # 计算热容的阈值（例如：超过平均值的10倍）
            mean_Cth = np.mean(cumulative_Cth)
            threshold = mean_Cth * 10

            # 找到需要保留的索引
            valid_indices = cumulative_Cth <= threshold

            if np.sum(valid_indices) < len(cumulative_Cth):
                print(f"去掉了 {len(cumulative_Cth) - np.sum(valid_indices)} 个热容过大的层")
                cumulative_Rth = cumulative_Rth[valid_indices]
                cumulative_Cth = cumulative_Cth[valid_indices]

                # 同时更新微分结构函数
                if len(differential_Rth) > 0:
                    # 确保微分结构函数的长度与积分结构函数匹配
                    max_valid_index = np.max(np.where(valid_indices)[0])
                    if max_valid_index < len(differential_Rth):
                        differential_Rth = differential_Rth[:max_valid_index]
                        differential_Cth = differential_Cth[:max_valid_index]

        # 保存结果 - 直接保存处理后的数组，不再使用valid_mask索引
        self.results['cumulative_Rth'] = cumulative_Rth
        self.results['cumulative_Cth'] = cumulative_Cth
        self.results['differential_Rth'] = differential_Rth
        self.results['differential_Cth'] = differential_Cth

        print(f"结构函数计算完成:")
        print(f"  积分结构函数数据点: {len(cumulative_Rth)}")
        print(f"  微分结构函数数据点: {len(differential_Rth)}")
        # if len(differential_Rth) > 0:
        #     print(f"  微分热阻范围: {differential_Rth.min():.2e} - {differential_Rth.max():.2e} K/W")
        #     print(f"  微分热容范围: {differential_Cth.min():.2e} - {differential_Cth.max():.2e} Ws/K")

        return True

    def calculate_structure_functions_alternative(self):
        """备选的结构函数计算方法 - 基于Cauer网络参数
        
        这是备选的结构函数计算方法，使用相同的Cauer网络参数但可能有不同的处理逻辑。
        计算逻辑与主要方法相同，基于Matlab代码实现。
        """
        if ('cauerRth' not in self.results or
                'cauerCth' not in self.results):
            print("警告: 没有Cauer网络参数，请先执行Foster到Cauer转换")
            return False

        # 确保转换为numpy数组并处理可能的非数值类型
        try:
            cauerRth = self.results['cauerRth']
            cauerCth = self.results['cauerCth']

            # 处理列表类型
            if isinstance(cauerRth, list):
                cauerRth_numeric = []
                for r_val in cauerRth:
                    try:
                        r_numeric = float(r_val)
                        cauerRth_numeric.append(r_numeric)
                    except (TypeError, ValueError):
                        continue
                cauerRth = np.array(cauerRth_numeric, dtype=self.dtype)
            else:
                cauerRth = np.array(cauerRth, dtype=self.dtype)

            if isinstance(cauerCth, list):
                cauerCth_numeric = []
                for c_val in cauerCth:
                    try:
                        c_numeric = float(c_val)
                        cauerCth_numeric.append(c_numeric)
                    except (TypeError, ValueError):
                        continue
                cauerCth = np.array(cauerCth_numeric, dtype=self.dtype)
            else:
                cauerCth = np.array(cauerCth, dtype=self.dtype)

        except (ValueError, TypeError) as e:
            print(f"警告: Cauer参数转换为数值失败: {e}")
            return False

        # 过滤掉零值、负值和无效值
        valid_mask = (cauerRth > 0) & (cauerCth > 0) & np.isfinite(cauerRth) & np.isfinite(cauerCth)
        if not np.any(valid_mask):
            print("警告: 没有有效的Cauer网络参数用于结构函数计算")
            return False

        cauerRth_valid = cauerRth[valid_mask]
        cauerCth_valid = cauerCth[valid_mask]
        # Cauer网络的结构函数计算 - 基于Matlab代码逻辑
        # 积分结构函数计算
        cumulative_Rth = np.cumsum(cauerRth_valid).astype(self.dtype)
        cumulative_Cth = np.cumsum(cauerCth_valid).astype(self.dtype)

        # 微分结构函数计算 - 按照Matlab代码逻辑
        # 计算相邻积分值的差值
        diff_cumulative_Rth = np.diff(cumulative_Rth)
        diff_cumulative_Cth = np.diff(cumulative_Cth)

        # 计算微分结构函数
        # 注意：避免除零错误
        valid_diff_mask = diff_cumulative_Rth > 0
        if np.any(valid_diff_mask):
            # 使用正确的数组长度
            differential_Cth = np.zeros(len(diff_cumulative_Rth))
            differential_Rth = np.zeros(len(diff_cumulative_Rth))

            # 对于有效差值，计算微分热容
            differential_Cth[valid_diff_mask] = diff_cumulative_Cth[valid_diff_mask] / diff_cumulative_Rth[
                valid_diff_mask]
            differential_Rth[valid_diff_mask] = cumulative_Rth[:-1][valid_diff_mask]

            # 过滤掉无效值
            valid_differential_mask = (differential_Rth > 0) & (differential_Cth > 0) & np.isfinite(
                differential_Rth) & np.isfinite(differential_Cth)
            if np.any(valid_differential_mask):
                differential_Rth = differential_Rth[valid_differential_mask]
                differential_Cth = differential_Cth[valid_differential_mask]
            else:
                # 如果没有有效微分值，使用原始Cauer参数作为备选
                differential_Rth = cauerRth_valid
                differential_Cth = cauerCth_valid
        else:
            # 如果没有有效差值，使用原始Cauer参数
            differential_Rth = cauerRth_valid
            differential_Cth = cauerCth_valid

        # 去掉最后几个热容特别大的层（如果存在）
        if len(cumulative_Cth) > 4:
            # 计算热容的阈值（例如：超过平均值的10倍）
            mean_Cth = np.mean(cumulative_Cth)
            threshold = mean_Cth * 10

            # 找到需要保留的索引
            valid_indices = cumulative_Cth <= threshold

            if np.sum(valid_indices) < len(cumulative_Cth):
                print(f"去掉了 {len(cumulative_Cth) - np.sum(valid_indices)} 个热容过大的层")
                cumulative_Rth = cumulative_Rth[valid_indices]
                cumulative_Cth = cumulative_Cth[valid_indices]

                # 同时更新微分结构函数
                if len(differential_Rth) > 0:
                    # 确保微分结构函数的长度与积分结构函数匹配
                    max_valid_index = np.max(np.where(valid_indices)[0])
                    if max_valid_index < len(differential_Rth):
                        differential_Rth = differential_Rth[:max_valid_index]
                        differential_Cth = differential_Cth[:max_valid_index]

        # 保存结果
        self.results['cumulative_Rth'] = cumulative_Rth
        self.results['cumulative_Cth'] = cumulative_Cth
        self.results['differential_Rth'] = differential_Rth
        self.results['differential_Cth'] = differential_Cth

        # print(f"基于Cauer网络的备选结构函数计算完成: {len(cumulative_Rth)} 个数据点")
        # print(f"热阻范围: {differential_Rth.min():.2e} - {differential_Rth.max():.2e} K/W")
        # print(f"积分热容范围: {cumulative_Cth.min():.2e} - {cumulative_Cth.max():.2e} Ws/K")
        # print(f"微分热容范围: {differential_Cth.min():.2e} - {differential_Cth.max():.2e} Ws/K")

        # 验证积分结构函数的阶梯状特性
        if len(cumulative_Cth) > 1:
            steps = np.diff(cumulative_Cth)
            # print(f"积分热容阶梯数: {len(steps)}")
            # print(f"阶梯高度范围: {steps.min():.2e} - {steps.max():.2e} Ws/K")
            # print(f"平均阶梯高度: {steps.mean():.2e} Ws/K")

        return True

    def full_analysis(self, file_path, ambient_temp=25.0):
        """执行完整分析流程"""
        self.results = {}

        if not self.load_data(file_path):
            return False

        if not self.calculate_zth(ambient_temp):
            return False

        if not self.logarithmic_interpolation():
            return False

        if not self.calculate_derivative():
            return False

        if not self.calculate_weight_function():
            return False

        if not self.bayesian_deconvolution():
            return False

        if not self.discrete_time_constant_spectrum():
            return False

        # self.manual_handler()

        # 尝试连分式展开法进行Foster到Cauer转换
        if not self.foster_to_cauer():
            return False

        # 基于Cauer网络计算结构函数
        if not self.calculate_structure_functions():
            # # 如果主要方法失败，尝试备选方法
            # if not self.calculate_structure_functions_alternative():
            #     print("警告: 结构函数计算失败")
            return False

        return True

    def analysis_from_time_constant_spectrum(self, file_path):
        """从时间常数谱数据开始执行分析流程"""
        self.results = {}

        # 直接加载时间常数谱数据
        if not self.load_time_constant_spectrum_data(file_path):
            return False

        # 跳过前面的计算步骤，直接从离散时间常数谱开始
        if not self.discrete_time_constant_spectrum():
            return False

        # 尝试连分式展开法进行Foster到Cauer转换
        if not self.foster_to_cauer():
            return False

        # 基于Cauer网络计算结构函数
        if not self.calculate_structure_functions():
            return False

        return True

    def export_transfer_function_results(self, file_path):
        """导出传递函数计算结果到Excel文件"""
        try:
            # 检查是否有可导出的结果
            exportable_results = {}
            
            # Foster网络参数
            if 'fosterRth' in self.results and 'fosterCth' in self.results:
                foster_rth = self.results['fosterRth']
                foster_cth = self.results['fosterCth']
                
                # 确保是numpy数组
                if isinstance(foster_rth, list):
                    foster_rth = np.array(foster_rth)
                if isinstance(foster_cth, list):
                    foster_cth = np.array(foster_cth)
                
                exportable_results['Foster_Rth_K_W'] = foster_rth
                exportable_results['Foster_Cth_Ws_K'] = foster_cth
                
                # 计算Foster网络的时间常数
                if len(foster_rth) == len(foster_cth):
                    foster_tau = foster_rth * foster_cth
                    exportable_results['Foster_Tau_s'] = foster_tau
            
            # Cauer网络参数
            if 'cauerRth' in self.results and 'cauerCth' in self.results:
                cauer_rth = self.results['cauerRth']
                cauer_cth = self.results['cauerCth']
                
                # 确保是numpy数组
                if isinstance(cauer_rth, list):
                    cauer_rth = np.array(cauer_rth)
                if isinstance(cauer_cth, list):
                    cauer_cth = np.array(cauer_cth)
                
                exportable_results['Cauer_Rth_K_W'] = cauer_rth
                exportable_results['Cauer_Cth_Ws_K'] = cauer_cth
                
                # 计算Cauer网络的时间常数
                if len(cauer_rth) == len(cauer_cth):
                    cauer_tau = cauer_rth * cauer_cth
                    exportable_results['Cauer_Tau_s'] = cauer_tau
            
            # 结构函数结果
            if 'cumulative_Rth' in self.results and 'cumulative_Cth' in self.results:
                exportable_results['Cumulative_Rth_K_W'] = self.results['cumulative_Rth']
                exportable_results['Cumulative_Cth_Ws_K'] = self.results['cumulative_Cth']
            
            if 'differential_Rth' in self.results and 'differential_Cth' in self.results:
                exportable_results['Differential_Rth_K_W'] = self.results['differential_Rth']
                exportable_results['Differential_Cth_Ws_K'] = self.results['differential_Cth']
            
            if not exportable_results:
                print("警告: 没有可导出的传递函数计算结果")
                return False
            
            # 创建DataFrame
            max_length = max(len(v) for v in exportable_results.values())
            
            # 填充较短的数组
            for key, value in exportable_results.items():
                if len(value) < max_length:
                    # 用NaN填充
                    padded_value = np.full(max_length, np.nan)
                    padded_value[:len(value)] = value
                    exportable_results[key] = padded_value
            
            df = pd.DataFrame(exportable_results)
            
            # 保存到Excel文件
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Transfer_Function_Results', index=False)
                
                # 添加元数据工作表
                metadata = {
                    'Parameter': ['Calculation_Date', 'Precision', 'Discrete_Order', 'Delta_z'],
                    'Value': [
                        pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                        self.precision,
                        self.discrete_order,
                        self.delta_z
                    ]
                }
                metadata_df = pd.DataFrame(metadata)
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
            
            print(f"传递函数计算结果已导出到: {file_path}")
            print(f"导出的参数数量: {len(exportable_results)}")
            print(f"数据点数量: {max_length}")
            return True
            
        except Exception as e:
            print(f"导出传递函数结果时出错: {e}")
            return False

    def import_transfer_function_results(self, file_path):
        """从Excel文件导入传递函数计算结果"""
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path, sheet_name='Transfer_Function_Results')
            
            # 导入Foster网络参数
            if 'Foster_Rth_K_W' in df.columns and 'Foster_Cth_Ws_K' in df.columns:
                foster_rth = df['Foster_Rth_K_W'].dropna().values.astype(self.dtype)
                foster_cth = df['Foster_Cth_Ws_K'].dropna().values.astype(self.dtype)
                
                if len(foster_rth) > 0 and len(foster_cth) > 0:
                    self.results['fosterRth'] = foster_rth
                    self.results['fosterCth'] = foster_cth
                    print(f"导入Foster网络参数: {len(foster_rth)} 个参数")
            
            # 导入Cauer网络参数
            if 'Cauer_Rth_K_W' in df.columns and 'Cauer_Cth_Ws_K' in df.columns:
                cauer_rth = df['Cauer_Rth_K_W'].dropna().values.astype(self.dtype)
                cauer_cth = df['Cauer_Cth_Ws_K'].dropna().values.astype(self.dtype)
                
                if len(cauer_rth) > 0 and len(cauer_cth) > 0:
                    self.results['cauerRth'] = cauer_rth
                    self.results['cauerCth'] = cauer_cth
                    print(f"导入Cauer网络参数: {len(cauer_rth)} 个参数")
            
            # 导入结构函数结果
            if 'Cumulative_Rth_K_W' in df.columns and 'Cumulative_Cth_Ws_K' in df.columns:
                cumulative_rth = df['Cumulative_Rth_K_W'].dropna().values.astype(self.dtype)
                cumulative_cth = df['Cumulative_Cth_Ws_K'].dropna().values.astype(self.dtype)
                
                if len(cumulative_rth) > 0 and len(cumulative_cth) > 0:
                    self.results['cumulative_Rth'] = cumulative_rth
                    self.results['cumulative_Cth'] = cumulative_cth
                    print(f"导入积分结构函数: {len(cumulative_rth)} 个数据点")
            
            if 'Differential_Rth_K_W' in df.columns and 'Differential_Cth_Ws_K' in df.columns:
                differential_rth = df['Differential_Rth_K_W'].dropna().values.astype(self.dtype)
                differential_cth = df['Differential_Cth_Ws_K'].dropna().values.astype(self.dtype)
                
                if len(differential_rth) > 0 and len(differential_cth) > 0:
                    self.results['differential_Rth'] = differential_rth
                    self.results['differential_Cth'] = differential_cth
                    print(f"导入微分结构函数: {len(differential_rth)} 个数据点")
            
            # 读取元数据
            try:
                metadata_df = pd.read_excel(file_path, sheet_name='Metadata')
                if 'Parameter' in metadata_df.columns and 'Value' in metadata_df.columns:
                    for _, row in metadata_df.iterrows():
                        param = row['Parameter']
                        value = row['Value']
                        if param == 'Precision' and value in ['float32', 'float64']:
                            self.precision = value
                            self.dtype = np.float32 if value == 'float32' else np.float64
                        elif param == 'Discrete_Order':
                            try:
                                self.discrete_order = int(value)
                            except:
                                pass
                        elif param == 'Delta_z':
                            try:
                                self.delta_z = float(value)
                            except:
                                pass
            except:
                print("警告: 无法读取元数据")
            
            print(f"传递函数计算结果已从文件导入: {file_path}")
            return True
            
        except Exception as e:
            print(f"导入传递函数结果时出错: {e}")
            return False
